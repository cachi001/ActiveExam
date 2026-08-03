"""Export Excel (.xlsx) del sumario de estadísticas institucionales (C-20).

openpyxl — hoja "Panel" con gráficos matplotlib embebidos (PNG, alta calidad)
más hojas de datos con tablas profesionales: encabezado azul, filas alternadas,
bordes completos, anchos auto-calculados, grillas ocultas. Sin PII. L2.5.
"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.application.stats.charts import dashboard_png
from app.application.stats.labels import etiqueta_evento
from app.application.stats.resumen_service import FiltrosStats, ResumenStats

# ── Paleta ───────────────────────────────────────────────────────────────────
_AZUL        = "004BA8"
_ZEBRA       = "F1F5F9"
_VERDE_BG    = "D1FAE5"
_ROJO_BG     = "FEE2E2"
_AMBAR_BG    = "FEF3C7"

_HDR_FILL  = PatternFill("solid", fgColor=_AZUL)
_HDR_FONT  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
_ZEBRA_FILL = PatternFill("solid", fgColor=_ZEBRA)

_THIN   = Side(style="thin",   color="CBD5E1")
_MEDIUM = Side(style="medium", color=_AZUL)
_BORDE  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDE_HDR = Border(left=_MEDIUM, right=_MEDIUM, top=_MEDIUM, bottom=_MEDIUM)

DECISION_LABEL = {
    "sin_revisar":      "Sin revisar",
    "pendiente":        "Pendiente",
    "aprobado":         "Aprobado",
    "anulado":          "Anulado por fraude",
}


# ── Helpers ──────────────────────────────────────────────────────────────────

def _filtros_txt(filtros: FiltrosStats | None, alcance: str | None) -> str:
    """Alcance del informe. ``alcance`` viene resuelto contra la base (nombres
    reales de materia/comisión/examen); el fallback solo cubre el caso sin filtros."""
    if alcance:
        return alcance
    activos = filtros and any([
        filtros.materia_id, filtros.comision_id,
        filtros.examen_contenido_id, filtros.desde, filtros.hasta,
    ])
    return "Filtrado" if activos else "Sin filtros — todo el período disponible"


def _auto_width(ws, col_idx: int, min_w: float = 10.0, max_w: float = 55.0) -> None:
    letter = get_column_letter(col_idx)
    max_len = 0
    for cell in ws[letter]:
        try:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        except Exception:
            pass
    ws.column_dimensions[letter].width = min(max(max_len + 3, min_w), max_w)


def _titulo_hoja(ws, titulo: str, subtitulo: str = "") -> None:
    """Escribe título + subtítulo en A1/A2 y oculta las grillas."""
    ws.sheet_view.showGridLines = False
    t = ws["A1"]
    t.value = titulo
    t.font  = Font(bold=True, size=13, color=_AZUL, name="Calibri")
    ws.row_dimensions[1].height = 22
    if subtitulo:
        s = ws["A2"]
        s.value = subtitulo
        s.font  = Font(italic=True, size=9, color="64748B", name="Calibri")
        ws.row_dimensions[2].height = 16


def _tabla(
    ws,
    fila_inicio: int,
    cols: list[str],
    filas: list[tuple],
    *,
    col_nums: set[int] | None = None,     # columnas (1-based) con alineación derecha
    fills: dict[int, str] | None = None,  # fila de dato (1-based) → color hex
) -> tuple[int, int]:
    """
    Escribe una tabla con encabezado azul + filas alternadas + bordes completos.
    Retorna (fila_encabezado, fila_ultimo_dato) para anclar gráficos.
    """
    col_nums = col_nums or set()
    fills    = fills    or {}

    # Encabezado
    for c, nombre in enumerate(cols, start=1):
        cell = ws.cell(row=fila_inicio, column=c, value=nombre)
        cell.fill      = _HDR_FILL
        cell.font      = _HDR_FONT
        cell.border    = _BORDE_HDR
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[fila_inicio].height = 20

    # Datos
    ultima = fila_inicio
    for i, fila in enumerate(filas):
        r = fila_inicio + 1 + i
        fill_hex = fills.get(i + 1)
        is_zebra = (i % 2 == 1) and not fill_hex
        for c, valor in enumerate(fila, start=1):
            cell = ws.cell(row=r, column=c, value=valor)
            cell.border = _BORDE
            cell.font   = Font(size=10, name="Calibri")
            if fill_hex:
                cell.fill = PatternFill("solid", fgColor=fill_hex)
            elif is_zebra:
                cell.fill = _ZEBRA_FILL
            is_num = isinstance(valor, (int, float)) or (c in col_nums)
            cell.alignment = Alignment(
                horizontal="right" if is_num else "left",
                vertical="center",
            )
        ws.row_dimensions[r].height = 18
        ultima = r

    return fila_inicio, ultima


def _grafico_barras(
    ws,
    hdr_row: int,
    ultima_row: int,
    col_cats: int,
    col_vals: int,
    titulo: str,
    ancla: str,
    *,
    first_data_row: int | None = None,
    horizontal: bool = False,
    alto: float = 12.0,
    ancho: float = 18.0,
) -> None:
    """first_data_row: fila de inicio de datos (default: hdr_row+1)."""
    fdr = first_data_row if first_data_row is not None else hdr_row + 1
    datos = Reference(ws, min_col=col_vals, min_row=hdr_row, max_row=ultima_row)
    cats  = Reference(ws, min_col=col_cats, min_row=fdr,     max_row=ultima_row)
    chart = BarChart()
    chart.type   = "bar" if horizontal else "col"
    chart.title  = titulo
    chart.legend = None
    chart.add_data(datos, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = alto
    chart.width  = ancho
    ws.add_chart(chart, ancla)


def _grafico_torta(
    ws,
    hdr_row: int,
    ultima_row: int,
    col_cats: int,
    col_vals: int,
    titulo: str,
    ancla: str,
    *,
    first_data_row: int | None = None,
    alto: float = 12.0,
    ancho: float = 18.0,
) -> None:
    """first_data_row: fila de inicio de datos (default: hdr_row+1)."""
    fdr = first_data_row if first_data_row is not None else hdr_row + 1
    datos = Reference(ws, min_col=col_vals, min_row=hdr_row, max_row=ultima_row)
    cats  = Reference(ws, min_col=col_cats, min_row=fdr,     max_row=ultima_row)
    chart = PieChart()
    chart.title = titulo
    chart.add_data(datos, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = alto
    chart.width  = ancho
    ws.add_chart(chart, ancla)


# ── Export principal ──────────────────────────────────────────────────────────

def resumen_a_xlsx(
    r: ResumenStats,
    filtros: FiltrosStats | None = None,
    alcance: str | None = None,
) -> bytes:
    """Serializa el sumario a un .xlsx profesional con tablas + gráficos.

    ``alcance``: descripción del recorte con los nombres ya resueltos contra la
    base (``describir_alcance``). El export no toca la DB.
    """
    wb = Workbook()

    # ── Hoja Panel ────────────────────────────────────────────────────────────
    # Imagen matplotlib de alta calidad — el chart más confiable en cualquier visor.
    panel = wb.active
    panel.title = "Panel"
    panel.sheet_view.showGridLines = False
    panel.sheet_view.zoomScale = 90

    panel["A1"] = "ActiveExam — Estadísticas institucionales"
    panel["A1"].font = Font(bold=True, size=15, color=_AZUL, name="Calibri")
    panel.row_dimensions[1].height = 26

    panel["A2"] = _filtros_txt(filtros, alcance)
    panel["A2"].font = Font(italic=True, size=10, color="64748B", name="Calibri")

    panel["A3"] = f"Generado: {datetime.now().strftime('%d/%m/%Y  %H:%M')}"
    panel["A3"].font = Font(size=9, color="94A3B8", name="Calibri")

    png_bytes = dashboard_png(r)
    img = XLImage(io.BytesIO(png_bytes))
    orig_w, orig_h = img.width, img.height
    img.width  = 920
    img.height = int(orig_h * (920 / orig_w)) if orig_w else 600
    panel.add_image(img, "A5")
    panel.column_dimensions["A"].width = 135

    # ── Hoja Resumen ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Resumen")
    _titulo_hoja(ws, "Resumen del período", _filtros_txt(filtros, alcance))

    metricas = [
        ("Exámenes",             r.total_examenes),
        ("Materias",             r.total_materias),
        ("Comisiones",           r.total_comisiones),
        ("Sesiones totales",     r.total_sesiones),
        ("Sesiones finalizadas", r.sesiones_finalizadas),
        (f"En riesgo  (score ≥ {r.umbral_riesgo})", r.sesiones_en_riesgo),
    ]
    hdr, ult = _tabla(ws, 4, ["Métrica", "Valor"], metricas, col_nums={2})
    _auto_width(ws, 1, min_w=35)
    _auto_width(ws, 2, min_w=12)

    # ── Hoja Habilitación ─────────────────────────────────────────────────────
    ws_h = wb.create_sheet("Habilitación")
    _titulo_hoja(
        ws_h,
        "Habilitación para rendir",
        "Requisito previo: consentimiento vigente + biometría de referencia",
    )
    el = r.elegibilidad
    hab_filas = [
        ("Total inscriptos",   el.total_inscriptos),
        ("Pueden rendir",      el.pueden_rendir),
        ("No pueden rendir",   el.no_pueden_rendir),
        ("Sin consentimiento", el.sin_consentimiento),
        ("Sin biometría",      el.sin_biometria),
    ]
    hdr_h, ult_h = _tabla(
        ws_h, 4, ["Estado", "Alumnos"], hab_filas,
        col_nums={2},
        fills={2: _VERDE_BG, 3: _ROJO_BG, 4: _AMBAR_BG, 5: _AMBAR_BG},
    )
    _auto_width(ws_h, 1, min_w=26)
    _auto_width(ws_h, 2, min_w=14)
    # El gráfico muestra solo Pueden / No pueden (filas 2-3 de la tabla = hdr_h+2..hdr_h+3).
    # "Total inscriptos" (fila 1) y los desglose (filas 4-5) quedan en la tabla pero no
    # distorsionan la torta — el par Pueden/No pueden ya suma el 100% del padrón.
    _grafico_torta(ws_h, hdr_h, hdr_h + 3, 1, 2, "Pueden rendir vs. No pueden rendir", "D4",
                   first_data_row=hdr_h + 2, alto=14, ancho=20)

    # ── Hoja Scores ───────────────────────────────────────────────────────────
    ws_sc = wb.create_sheet("Scores")
    _titulo_hoja(ws_sc, "Distribución de scores")
    # Las bandas ya vienen ordenadas de menor a mayor desde el servicio, con la
    # última arrancando en el umbral vivo — no se asume ninguna etiqueta fija.
    dist_filas = list(r.distribucion_scores.items())
    hdr_sc, ult_sc = _tabla(ws_sc, 3, ["Rango de score", "Sesiones"], dist_filas, col_nums={2})
    _auto_width(ws_sc, 1, min_w=20)
    _auto_width(ws_sc, 2, min_w=14)
    _grafico_barras(ws_sc, hdr_sc, ult_sc, 1, 2, "Distribución de scores", "D3",
                    alto=14, ancho=20)

    # ── Hoja Por materia ──────────────────────────────────────────────────────
    ws_m = wb.create_sheet("Por materia")
    _titulo_hoja(ws_m, "Sesiones por materia")
    mat_filas = [(m.nombre, m.sesiones, m.en_riesgo) for m in (r.por_materia or [])]
    if mat_filas:
        hdr_m, ult_m = _tabla(ws_m, 3, ["Materia", "Sesiones", "En riesgo"], mat_filas, col_nums={2, 3})
        _auto_width(ws_m, 1, min_w=32)
        _auto_width(ws_m, 2, min_w=12)
        _auto_width(ws_m, 3, min_w=12)
        _grafico_barras(ws_m, hdr_m, ult_m, 1, 2, "Sesiones por materia", "E3",
                        horizontal=True, alto=max(10, len(mat_filas) * 1.4), ancho=20)
    else:
        ws_m["A3"] = "Sin datos"

    # ── Hoja Por comisión ─────────────────────────────────────────────────────
    ws_c = wb.create_sheet("Por comisión")
    _titulo_hoja(ws_c, "Sesiones por comisión")
    com_filas = [(c.nombre, c.sesiones, c.en_riesgo) for c in (r.por_comision or [])]
    if com_filas:
        hdr_c, ult_c = _tabla(ws_c, 3, ["Comisión", "Sesiones", "En riesgo"], com_filas, col_nums={2, 3})
        _auto_width(ws_c, 1, min_w=32)
        _auto_width(ws_c, 2, min_w=12)
        _auto_width(ws_c, 3, min_w=12)
        _grafico_barras(ws_c, hdr_c, ult_c, 1, 2, "Sesiones por comisión", "E3",
                        horizontal=True, alto=max(10, len(com_filas) * 1.4), ancho=20)
    else:
        ws_c["A3"] = "Sin datos"

    # ── Hoja Detectores ───────────────────────────────────────────────────────
    ws_d = wb.create_sheet("Detectores")
    _titulo_hoja(ws_d, "Detectores más frecuentes")
    det_filas = [(etiqueta_evento(e.tipo), e.cantidad) for e in (r.top_eventos or [])]
    if det_filas:
        hdr_d, ult_d = _tabla(ws_d, 3, ["Detector", "Cantidad"], det_filas, col_nums={2})
        _auto_width(ws_d, 1, min_w=30)
        _auto_width(ws_d, 2, min_w=14)
        _grafico_barras(ws_d, hdr_d, ult_d, 1, 2, "Detectores más frecuentes", "D3",
                        horizontal=True, alto=max(10, len(det_filas) * 1.4), ancho=20)
    else:
        ws_d["A3"] = "Sin datos"

    # ── Hoja Estado de revisión ───────────────────────────────────────────────
    ws_r = wb.create_sheet("Revisión")
    _titulo_hoja(ws_r, "Estado de revisión")
    dec_items = sorted((r.decisiones or {}).items(), key=lambda kv: -kv[1])
    rev_filas = [(DECISION_LABEL.get(k, k), v) for k, v in dec_items]
    if rev_filas:
        hdr_r, ult_r = _tabla(ws_r, 3, ["Estado", "Sesiones"], rev_filas, col_nums={2})
        _auto_width(ws_r, 1, min_w=26)
        _auto_width(ws_r, 2, min_w=14)
        _grafico_torta(ws_r, hdr_r, ult_r, 1, 2, "Estado de revisión", "D3",
                       alto=14, ancho=20)
    else:
        ws_r["A3"] = "Sin datos"

    # ── Hoja Actividad ────────────────────────────────────────────────────────
    ws_a = wb.create_sheet("Actividad")
    _titulo_hoja(ws_a, "Actividad por día")
    act_filas = [(d.fecha, d.sesiones) for d in (r.por_dia or [])]
    if act_filas:
        hdr_a, ult_a = _tabla(ws_a, 3, ["Fecha", "Sesiones"], act_filas, col_nums={2})
        _auto_width(ws_a, 1, min_w=16)
        _auto_width(ws_a, 2, min_w=14)
        _grafico_barras(ws_a, hdr_a, ult_a, 1, 2, "Actividad por día", "D3",
                        alto=12, ancho=20)
    else:
        ws_a["A3"] = "Sin datos"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
