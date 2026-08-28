"""Export de listados académicos a Excel y PDF (c-78 §13, E-10).

Existe porque hay campus SIN API: la nota se termina cargando a mano, y para eso
hace falta el listado en un archivo, no en una pantalla paginada. Lo mismo para
cruzar los inscriptos de una comisión contra el padrón del campus.

Dos exportables, un solo motor:
  - inscriptos de una comisión (apellido, nombre, usuario, email, inscripción)
  - notas de un examen (alumno, nota, estado de entrega, estado en el campus)

El motor (`tabla_a_xlsx` / `tabla_a_pdf`) es genérico a propósito: los dos
exportables difieren solo en columnas y encabezado, y duplicar el armado del
Workbook/FPDF garantiza que uno de los dos se quede atrás. Mismo criterio de
estilo que `audit/export.py`, del que se toma el patrón ya validado (incluida la
transliteración a latin-1 que FPDF necesita).

PRIVACIDAD (Ley 25.326): estos archivos llevan datos personales de alumnos. Se
exporta lo MÍNIMO para el propósito declarado — identificar a la persona en el
campus. No van scores de proctoring, ni eventos, ni evidencia.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.domain.exam_content.estado_entrega import etiqueta_estado_entrega
from app.domain.exam_content.resultado_nota import etiqueta_resultado, etiqueta_retencion

_AZUL = "004BA8"
_HDR_FILL = PatternFill("solid", fgColor=_AZUL)
_HDR_FONT = Font(bold=True, color="FFFFFF")

# Ver el comentario extenso en audit/export.py: FPDF con fuente core solo soporta
# latin-1, y las rayas/comillas tipográficas quedaban como "?".
_REEMPLAZOS_PDF_LATIN1 = {
    "—": "-",
    "–": "-",
    "…": "...",
    "→": "->",
    "«": '"',
    "»": '"',
    "‘": "'",
    "’": "'",
    "“": '"',
    "”": '"',
    "·": "-",
}


def _txt(valor: str) -> str:
    """Translitera lo tipográfico y reemplaza lo que aun así no entra en latin-1."""
    texto = valor or ""
    for feo, equivalente in _REEMPLAZOS_PDF_LATIN1.items():
        texto = texto.replace(feo, equivalente)
    return texto.encode("latin-1", "replace").decode("latin-1")


def fecha_legible(valor) -> str:
    """datetime/ISO → 'dd/mm/aaaa hh:mm'. Vacío si no hay dato (nunca 'None')."""
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y %H:%M")
    try:
        return datetime.fromisoformat(str(valor).replace("Z", "+00:00")).strftime(
            "%d/%m/%Y %H:%M"
        )
    except (ValueError, TypeError):
        return str(valor)


@dataclass(frozen=True, slots=True)
class Columna:
    """Una columna del export: su título y su ancho en cada formato."""

    titulo: str
    #: Ancho en "caracteres" de openpyxl.
    ancho_xlsx: int
    #: Ancho en mm para FPDF.
    ancho_pdf: int


#: Un valor negativo se pinta en rojo y uno afirmativo en verde. La regla vive acá
#: y no en cada export porque el criterio es el mismo en todos: lo que bloquea al
#: alumno tiene que saltar a la vista sin leer la fila entera.
_VERDE_FILL = PatternFill("solid", fgColor="E3F5E9")
_VERDE_FONT = Font(color="1B5E20", bold=True)
_ROJO_FILL = PatternFill("solid", fgColor="FDE7E9")
_ROJO_FONT = Font(color="A31515", bold=True)


def _tono_de(valor: str) -> str | None:
    """'ok', 'mal' o None según lo que dice la celda. None = sin color."""
    v = valor.strip().lower()
    if v == "sí" or v == "si":
        return "ok"
    if v == "no" or v.startswith("no —") or v.startswith("no -"):
        return "mal"
    return None


@dataclass(frozen=True, slots=True)
class Metrica:
    """Un número del encabezado, como tarjeta.

    El tono no es decoración: es lo que hace que el número se entienda sin leer
    la etiqueta. Un cero en "faltan ambas" es una buena noticia y un cero en
    "pueden rendir" es una catástrofe; pintarlos igual obliga a leer todo.
    """

    etiqueta: str
    valor: str
    #: 'ok' | 'malo' | 'neutro'
    tono: str = "neutro"
    #: Línea chica bajo el número ("de 9"). Un "1" suelto no se puede leer.
    detalle: str = ""


#: Colores de las tarjetas, por tono. (fondo, texto) en RGB para el PDF y en hex
#: para el Excel: son los mismos, escritos en los dos formatos que cada librería
#: necesita.
_TONOS_PDF = {
    "ok": ((227, 245, 233), (27, 94, 32)),
    "malo": ((253, 231, 233), (163, 21, 21)),
    "neutro": ((240, 243, 248), (0, 75, 168)),
}
_TONOS_XLSX = {
    "ok": ("E3F5E9", "1B5E20"),
    "malo": ("FDE7E9", "A31515"),
    "neutro": ("F0F3F8", "004BA8"),
}


def _tarjetas_pdf(pdf: FPDF, metricas: list[Metrica], ancho_util: int) -> None:
    """Dibuja las métricas como tarjetas en una fila.

    Se reparten el ancho útil en partes iguales. Si son muchas quedan angostas,
    pero es preferible a que una se caiga del papel: FPDF no avisa cuando algo
    se dibuja fuera de la página.
    """
    if not metricas:
        return
    separacion = 3
    ancho = (ancho_util - separacion * (len(metricas) - 1)) / len(metricas)
    alto = 18
    x0, y0 = pdf.get_x(), pdf.get_y()

    for i, m in enumerate(metricas):
        fondo, tinta = _TONOS_PDF.get(m.tono, _TONOS_PDF["neutro"])
        x = x0 + i * (ancho + separacion)
        pdf.set_fill_color(*fondo)
        pdf.rect(x, y0, ancho, alto, style="F")

        # Etiqueta chica en mayúsculas arriba, número grande abajo: el orden del
        # informe de tutoría, que es el que se lee de un vistazo.
        pdf.set_xy(x + 2, y0 + 2)
        pdf.set_font("Helvetica", "B", 6.5)
        pdf.set_text_color(110, 110, 110)
        pdf.cell(ancho - 4, 3.5, _txt(m.etiqueta.upper()), align="L")

        pdf.set_xy(x + 2, y0 + 6)
        pdf.set_font("Helvetica", "B", 15)
        pdf.set_text_color(*tinta)
        pdf.cell(ancho - 4, 8, _txt(m.valor), align="L")

        if m.detalle:
            pdf.set_xy(x + 2, y0 + 13.5)
            pdf.set_font("Helvetica", "", 6.5)
            pdf.set_text_color(110, 110, 110)
            pdf.cell(ancho - 4, 3.5, _txt(m.detalle), align="L")

    pdf.set_xy(x0, y0 + alto)


def _criterios_pdf(pdf: FPDF, criterios: list[tuple[str, str]], ancho_util: int) -> None:
    """La leyenda que explica qué significa cada estado.

    Sin esto el lector tiene que deducir la diferencia entre dos columnas que
    ambas dicen "estado", que es exactamente lo que pasaba con "Entrega" y
    "Estado en el campus".
    """
    if not criterios:
        return
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 7)
    pdf.set_text_color(110, 110, 110)
    pdf.cell(0, 4, _txt("CÓMO LEER ESTE LISTADO"), new_x="LMARGIN", new_y="NEXT")
    for termino, explicacion in criterios:
        pdf.set_font("Helvetica", "B", 7.5)
        pdf.set_text_color(0, 75, 168)
        ancho_termino = pdf.get_string_width(_txt(termino)) + 2
        pdf.cell(ancho_termino, 4, _txt(termino))
        pdf.set_font("Helvetica", "", 7.5)
        pdf.set_text_color(90, 90, 90)
        pdf.multi_cell(
            ancho_util - ancho_termino, 4, _txt(explicacion), new_x="LMARGIN", new_y="NEXT"
        )


def tabla_a_xlsx(
    *,
    titulo: str,
    subtitulo: str,
    columnas: list[Columna],
    filas: list[list[str]],
    nombre_hoja: str = "Listado",
    metricas: list[Metrica] | None = None,
    criterios: list[tuple[str, str]] | None = None,
    colorear: bool = False,
) -> bytes:
    """Tabla genérica como .xlsx, con cabecera de contexto. Devuelve los bytes.

    La cabecera (título, subtítulo, fecha de generación y cantidad de filas) no
    es decoración: un listado suelto sin decir de qué comisión y de cuándo es no
    se puede usar para cruzar nada.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja[:31] or "Listado"

    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=14, color=_AZUL)
    ws["A2"] = subtitulo
    ws["A3"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A4"] = f"Filas: {len(filas)}"

    # Tarjetas ARRIBA de la tabla: es lo que se lee primero al abrir el archivo.
    # Sin esto habia que contar a mano quien se queda afuera, y con 80 inscriptos
    # eso no se hace. Cada metrica ocupa dos filas (etiqueta y numero) en su
    # propia columna, que es lo mas parecido a una tarjeta que da una planilla.
    fila_encabezado = 6
    if metricas:
        for i, m in enumerate(metricas, start=1):
            fondo, tinta = _TONOS_XLSX.get(m.tono, _TONOS_XLSX["neutro"])
            etiqueta = ws.cell(row=fila_encabezado, column=i, value=m.etiqueta.upper())
            etiqueta.font = Font(bold=True, size=8, color="6E6E6E")
            etiqueta.fill = PatternFill("solid", fgColor=fondo)
            etiqueta.alignment = Alignment(horizontal="left")

            texto = m.valor if not m.detalle else f"{m.valor} {m.detalle}"
            numero = ws.cell(row=fila_encabezado + 1, column=i, value=texto)
            numero.font = Font(bold=True, size=16, color=tinta)
            numero.fill = PatternFill("solid", fgColor=fondo)
            numero.alignment = Alignment(horizontal="left")
        ws.row_dimensions[fila_encabezado + 1].height = 24
        fila_encabezado += 3

    if criterios:
        titulo_criterios = ws.cell(
            row=fila_encabezado, column=1, value="CÓMO LEER ESTE LISTADO"
        )
        titulo_criterios.font = Font(bold=True, size=8, color="6E6E6E")
        fila_encabezado += 1
        for termino, explicacion in criterios:
            c_termino = ws.cell(row=fila_encabezado, column=1, value=termino)
            c_termino.font = Font(bold=True, size=9, color=_AZUL)
            c_exp = ws.cell(row=fila_encabezado, column=2, value=explicacion)
            c_exp.font = Font(size=9, color="5A5A5A")
            fila_encabezado += 1
        fila_encabezado += 1

    for i, col in enumerate(columnas, start=1):
        celda = ws.cell(row=fila_encabezado, column=i, value=col.titulo)
        celda.fill = _HDR_FILL
        celda.font = _HDR_FONT
        celda.alignment = Alignment(horizontal="left")
        ws.column_dimensions[chr(64 + i)].width = col.ancho_xlsx

    for j, fila in enumerate(filas, start=fila_encabezado + 1):
        for i, valor in enumerate(fila, start=1):
            celda = ws.cell(row=j, column=i, value=valor)
            if not colorear:
                continue
            tono = _tono_de(valor)
            if tono == "ok":
                celda.fill, celda.font = _VERDE_FILL, _VERDE_FONT
            elif tono == "mal":
                celda.fill, celda.font = _ROJO_FILL, _ROJO_FONT

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


#: Margen lateral que aplica FPDF por defecto (10 mm de cada lado).
_MARGEN_LATERAL_MM = 10
#: Ancho de página A4 en mm, en cada orientación.
_A4_ANCHO_MM = {False: 210, True: 297}


def ancho_util_mm(apaisado: bool) -> int:
    """Milímetros disponibles para la tabla, descontando los márgenes.

    Las celdas se dibujan una al lado de la otra con ancho fijo: lo que se pasa
    de este número queda FUERA del papel, sin error y sin aviso. Por eso el ancho
    de las columnas se valida contra este valor (`test_export_pdf_columnas_entran`).
    """
    return _A4_ANCHO_MM[bool(apaisado)] - 2 * _MARGEN_LATERAL_MM


def tabla_a_pdf(
    *,
    titulo: str,
    subtitulo: str,
    columnas: list[Columna],
    filas: list[list[str]],
    apaisado: bool = False,
    metricas: list[Metrica] | None = None,
    criterios: list[tuple[str, str]] | None = None,
    colorear: bool = False,
) -> bytes:
    """Tabla genérica como PDF. Devuelve los bytes.

    El PDF es una TABLA para imprimir/mirar: los textos largos se recortan al
    ancho de su columna. El archivo completo sin recortes es el Excel (mismo
    criterio que el export de auditoría).
    """
    pdf = FPDF(orientation="L" if apaisado else "P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 15)
    pdf.set_text_color(0, 75, 168)
    pdf.cell(0, 9, _txt(titulo), new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 5, _txt(subtitulo), new_x="LMARGIN", new_y="NEXT")
    pdf.cell(
        0,
        5,
        _txt(
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} - "
            f"Filas: {len(filas)}"
        ),
        new_x="LMARGIN",
        new_y="NEXT",
    )

    # Tarjetas de metricas ARRIBA de la tabla. Como el informe de tutoria de la
    # catedra: etiqueta chica en mayusculas, numero grande, y el color diciendo si
    # el numero es bueno o malo. Un renglon de texto corrido se pierde entre el
    # titulo y la tabla; una tarjeta se lee de un vistazo, que es para lo que
    # existe un resumen.
    if metricas:
        pdf.ln(3)
        _tarjetas_pdf(pdf, metricas, ancho_util_mm(apaisado))
    _criterios_pdf(pdf, criterios or [], ancho_util_mm(apaisado))
    pdf.ln(3)

    anchos = [c.ancho_pdf for c in columnas]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(0, 75, 168)
    pdf.set_text_color(255, 255, 255)
    for ancho, col in zip(anchos, columnas):
        pdf.cell(ancho, 8, _txt(col.titulo), border=0, fill=True, align="L")
    pdf.ln(8)

    pdf.set_font("Helvetica", "", 8)
    for i, fila in enumerate(filas):
        base = (245, 247, 250) if i % 2 == 0 else (255, 255, 255)
        for ancho, valor in zip(anchos, fila):
            maximo = max(4, int(ancho / 1.7))
            texto = valor if len(valor) <= maximo else valor[: maximo - 3] + "..."
            # La celda que dice si el alumno puede rendir se pinta: en una hoja de
            # 80 renglones de "Sí/No" el color es lo único que deja ver de un
            # vistazo quién se queda afuera.
            tono = _tono_de(valor) if colorear else None
            if tono == "ok":
                pdf.set_fill_color(227, 245, 233)
                pdf.set_text_color(27, 94, 32)
            elif tono == "mal":
                pdf.set_fill_color(253, 231, 233)
                pdf.set_text_color(163, 21, 21)
            else:
                pdf.set_fill_color(*base)
                pdf.set_text_color(30, 30, 30)
            pdf.cell(ancho, 6, _txt(texto), border=0, fill=True, align="L")
        pdf.ln(6)

    salida = pdf.output()
    return bytes(salida) if not isinstance(salida, bytes) else salida


# ---------------------------------------------------------------------------
# Inscriptos de una comisión (E-10)
# ---------------------------------------------------------------------------

# Las columnas para cruzar contra el padrón de Moodle, MÁS la elegibilidad de cada
# alumno. Sin consentimiento o sin biometría el alumno no puede rendir, y ese es el
# motivo por el que se descarga este listado antes del examen: saber a quién hay que
# avisarle. La pantalla ya lo mostraba por alumno; el archivo lo omitía, así que
# había que revisar uno por uno en la web lo que el export debía resolver de una.
#: El PDF de inscriptos va APAISADO. Con las 8 columnas, en A4 vertical entran
#: 186 mm y la tabla necesita más: las tres de elegibilidad quedaban fuera del
#: papel, así que en el teléfono el archivo se cortaba en "Inscripción" y no se
#: veía justamente el dato por el que se abre el listado.
INSCRIPTOS_APAISADO = True

COLUMNAS_INSCRIPTOS = [
    Columna("Apellido", 26, 26),
    Columna("Nombre", 26, 26),
    Columna("Usuario", 26, 24),
    # 50 mm entra un mail institucional completo (juan.perez@frm.utn.edu.ar).
    # Uno más largo se recorta: el archivo sin recortes es el Excel.
    Columna("Email", 38, 50),
    # 30 mm y no menos: con 26 la fecha salía como "27/08/2026 1...".
    Columna("Inscripción", 20, 30),
    Columna("Consentimiento", 16, 24),
    Columna("Biometría", 14, 20),
    # Va última y con el motivo adentro: es la columna que se lee primero cuando
    # se abre el archivo para ver quién se queda afuera. Se le deja todo el ancho
    # que sobra porque su texto es el más largo ("NO — Falta consentimiento y
    # biometría") y recortarlo deja el motivo a medias, que es peor que no tenerlo.
    Columna("¿Puede rendir?", 34, 72),
]


#: Leyenda del listado de inscriptos. Explica qué habilita a rendir y por qué,
#: que es la pregunta que el archivo viene a responder.
CRITERIOS_INSCRIPTOS = [
    (
        "Puede rendir:",
        "necesita las dos cosas, el consentimiento firmado y la captura biométrica. "
        "Con una sola no alcanza.",
    ),
    (
        "Consentimiento:",
        "lo firma el alumno desde su cuenta, sin cámara. Se puede resolver a distancia.",
    ),
    (
        "Biometría:",
        "es la captura de rostro de referencia. Necesita cámara y buena luz.",
    ),
]

#: Leyenda del listado de notas. Existe porque las dos últimas columnas se leían
#: como lo mismo: ambas decían "estado" al lado de "Entregada" y "Cargada".
CRITERIOS_NOTAS = [
    (
        "Resultado:",
        "si aprobó o no, comparando su nota con la nota de aprobación del examen. "
        "«Anulada» es una decisión humana: la nota que queda es 0 y al lado se "
        "muestra de cuánto venía.",
    ),
    (
        "Estado de la entrega:",
        "una tercera: si la nota ya se entregó al campus. Una nota calculada acá no "
        "está en la libreta hasta que diga «Cargada y confirmada» o «Cargada a mano».",
    ),
    (
        "Sin nota:",
        "todavía no hay nota. No es un cero ni un desaprobado.",
    ),
]


def _celda_nota(efectiva: object, calculada: object, retenido: object) -> str:
    """La nota que vale, y entre paréntesis la calculada si la anulación la cambió.

    Una celda vacía se lee como un error del archivo; decir que todavía no hay
    nota es un dato. Y en una anulación el 0 solo no alcanza: sin ver de cuánto
    venía, no se puede reclamar ni auditar la decisión.
    """
    if efectiva is None:
        return "Sin nota"
    texto = f"{float(efectiva):.2f}"
    if retenido == "anulada" and calculada is not None:
        texto += f" (calculada: {float(calculada):.2f})"
    return texto


def _si_no(valor: object) -> str:
    """Sí/No legible. Un campo ausente se informa como No, nunca como Sí: decir
    que sí sin dato haría creer que el alumno está listo cuando no se sabe."""
    return "Sí" if valor is True else "No"


def _puede_rendir_legible(inscripto) -> str:
    """"Sí" o "NO — {motivo}". Sin motivo, "NO" a secas: la ausencia de razón no
    puede volverse un "sí" por descarte."""
    if getattr(inscripto, "puede_rendir", False) is True:
        return "Sí"
    razon = (getattr(inscripto, "razon", None) or "").strip()
    return f"NO — {razon}" if razon else "NO"


@dataclass(frozen=True)
class ResumenElegibilidad:
    """Cuántos pueden rendir y, de los que no, qué les falta.

    Las tres categorías de falta son EXCLUYENTES: quien no tiene ninguna de las
    dos cuenta una sola vez, en `faltan_ambas`. Si se solaparan, la suma daría
    más que el total y el número dejaría de servir para decidir si el examen se
    puede tomar.

    La distinción importa porque se resuelven distinto: el consentimiento lo
    firma el alumno desde su casa, la captura biométrica necesita cámara.
    """

    total: int
    pueden_rendir: int
    falta_consentimiento: int
    falta_biometria: int
    faltan_ambas: int

    @property
    def no_pueden_rendir(self) -> int:
        return self.total - self.pueden_rendir

    def lineas(self) -> list[str]:
        """Las líneas del bloque de resumen del archivo.

        Las faltas en cero NO se imprimen: con el curso entero listo, un
        "Falta consentimiento: 0" es ruido que compite con el único número que
        importa.
        """
        salida = [f"Pueden rendir: {self.pueden_rendir} de {self.total}"]
        if self.falta_consentimiento:
            salida.append(f"Falta consentimiento: {self.falta_consentimiento}")
        if self.falta_biometria:
            salida.append(f"Falta biometría: {self.falta_biometria}")
        if self.faltan_ambas:
            salida.append(f"Faltan consentimiento y biometría: {self.faltan_ambas}")
        return salida

    def metricas(self) -> list[Metrica]:
        """Las tarjetas del encabezado. Las faltas en cero no generan tarjeta:
        una que diga 0 compite por atención con las que importan."""
        tarjetas = [
            Metrica(
                etiqueta="Pueden rendir",
                valor=str(self.pueden_rendir),
                # Con uno solo que no pueda hay un problema el día del examen: el
                # número se ve como problema, no como un dato más.
                tono="ok" if self.no_pueden_rendir == 0 else "malo",
                detalle=f"de {self.total}",
            )
        ]
        if self.falta_consentimiento:
            tarjetas.append(
                Metrica("Falta consentimiento", str(self.falta_consentimiento), "malo")
            )
        if self.falta_biometria:
            tarjetas.append(Metrica("Falta biometría", str(self.falta_biometria), "malo"))
        if self.faltan_ambas:
            tarjetas.append(Metrica("Faltan las dos", str(self.faltan_ambas), "malo"))
        return tarjetas


def resumen_elegibilidad(inscriptos: list) -> ResumenElegibilidad:
    """Cuenta el padrón por estado de habilitación (función PURA).

    Un dato ausente cuenta como falta, nunca como listo: decir que sí sin saberlo
    manda a rendir a alguien que después no va a poder, y el error aparece el día
    del examen.
    """
    pueden = falta_c = falta_b = faltan_ambas = 0
    for i in inscriptos:
        con_consentimiento = getattr(i, "consentimiento_vigente", None) is True
        con_biometria = getattr(i, "biometria_vigente", None) is True
        if con_consentimiento and con_biometria:
            pueden += 1
        elif not con_consentimiento and not con_biometria:
            faltan_ambas += 1
        elif not con_consentimiento:
            falta_c += 1
        else:
            falta_b += 1
    return ResumenElegibilidad(
        total=len(inscriptos),
        pueden_rendir=pueden,
        falta_consentimiento=falta_c,
        falta_biometria=falta_b,
        faltan_ambas=faltan_ambas,
    )


def filas_inscriptos(inscriptos: list) -> list[list[str]]:
    """Proyecta los inscriptos a filas de export (función PURA)."""
    return [
        [
            getattr(i, "apellido", "") or "",
            getattr(i, "nombre", "") or "",
            getattr(i, "username", "") or "",
            getattr(i, "email", "") or "",
            fecha_legible(getattr(i, "inscripto_en", None)),
            _si_no(getattr(i, "consentimiento_vigente", None)),
            _si_no(getattr(i, "biometria_vigente", None)),
            _puede_rendir_legible(i),
        ]
        for i in inscriptos
    ]


# ---------------------------------------------------------------------------
# Notas de un examen (E-10)
# ---------------------------------------------------------------------------

#: También apaisado: seis columnas con emails y etiquetas largas no entran en
#: vertical. La orientación vive acá, al lado de los anchos que la determinan,
#: para que agregar una columna y validar que entre sea una sola decisión.
NOTAS_APAISADO = True

# Tres cosas distintas que se llamaban todas "estado", y el titulo ahora las
# separa: si el alumno APROBO (la nota), si TERMINO el examen (la entrega), y si
# la nota LLEGO al campus (el envio).
#
# El veredicto academico faltaba por completo: el archivo traia la nota cruda y
# habia que saberse de memoria con cuanto se aprueba para leerlo.
COLUMNAS_NOTAS = [
    Columna("Alumno", 30, 45),
    Columna("Usuario", 24, 36),
    Columna("Email", 34, 50),
    Columna("Nota", 10, 18),
    Columna("Resultado", 20, 30),
    Columna("Estado de la entrega", 28, 46),
]

#: Del EXAMEN, no de la nota: si el alumno lo terminó y si ya se corrigió.
_ETIQUETA_ENTREGA = {
    "no_finalizada": "No lo terminó",
    "en_revision": "En revisión",
    "revisada": "Revisada",
    "finalizada": "Entregada",
}


#: Estados en los que la nota YA está en la libreta del campus. El resto es
#: trabajo pendiente, incluido 'fallido': la nota existe, el alumno la ve, y el
#: campus no la tiene.
_ESTADOS_YA_EN_EL_CAMPUS = frozenset({"enviado", "manual"})


@dataclass(frozen=True)
class ResumenNotas:
    """Los dos números por los que se baja este archivo: cuántos aprobaron y
    cuántas notas todavía no llegaron al campus.

    "Sin nota" se cuenta aparte de "desaprobado": meter en el mismo bolsón a
    quien sacó 3 y a quien no rindió infla el número de desaprobados, y ese
    número se informa.
    """

    total: int
    aprobados: int
    desaprobados: int
    sin_nota: int
    sin_cargar: int

    def lineas(self) -> list[str]:
        salida = [f"Resultados: {self.total}"]
        if self.aprobados or self.desaprobados:
            salida[0] = (
                f"Aprobados: {self.aprobados} de {self.total}"
                f" · Desaprobados: {self.desaprobados}"
            )
        if self.sin_nota:
            salida.append(f"Sin nota todavía: {self.sin_nota}")
        if self.sin_cargar:
            salida.append(f"Faltan cargar en el campus: {self.sin_cargar}")
        return salida

    def metricas(self) -> list[Metrica]:
        tarjetas: list[Metrica] = []
        if self.aprobados or self.desaprobados:
            tarjetas.append(
                Metrica("Aprobados", str(self.aprobados), "ok", f"de {self.total}")
            )
            tarjetas.append(Metrica("Desaprobados", str(self.desaprobados), "neutro"))
        else:
            # Sin nota de aprobación no se puede decir quién aprobó. Mostrar
            # "Aprobados: 0" sería afirmar que nadie aprobó, que es distinto de
            # no saberlo. Se informa lo que sí se sabe.
            tarjetas.append(Metrica("Resultados", str(self.total), "neutro"))
        if self.sin_nota:
            tarjetas.append(Metrica("Sin nota todavía", str(self.sin_nota), "neutro"))
        if self.sin_cargar:
            tarjetas.append(Metrica("Faltan en Moodle", str(self.sin_cargar), "malo"))
        return tarjetas


def resumen_notas(resultados: list, *, nota_aprobacion: float | None) -> ResumenNotas:
    """Cuenta los resultados de un examen (función PURA).

    Sin `nota_aprobacion` utilizable no se inventa quién aprobó: se informan el
    total y lo que falta cargar, que no dependen del umbral.
    """
    hay_umbral = bool(nota_aprobacion)
    aprobados = desaprobados = sin_nota = sin_cargar = 0
    for r in resultados:
        nota = getattr(r, "nota", None)
        if nota is None:
            sin_nota += 1
        elif hay_umbral:
            if float(nota) >= float(nota_aprobacion):  # type: ignore[arg-type]
                aprobados += 1
            else:
                desaprobados += 1
        estado = getattr(r, "estado_moodle", "") or ""
        if estado not in _ESTADOS_YA_EN_EL_CAMPUS:
            sin_cargar += 1
    return ResumenNotas(
        total=len(resultados),
        aprobados=aprobados,
        desaprobados=desaprobados,
        sin_nota=sin_nota,
        sin_cargar=sin_cargar,
    )


def filas_notas(resultados: list) -> list[list[str]]:
    """Proyecta los resultados de un examen a filas de export (función PURA)."""
    filas: list[list[str]] = []
    for r in resultados:
        # `nota_efectiva` y no `nota`: una anulación deja la nota en 0, y el
        # archivo tiene que decir lo mismo que la pantalla. La calculada va al
        # lado, entre paréntesis, para no perder el dato.
        nota = getattr(r, "nota_efectiva", None)
        if nota is None:
            nota = getattr(r, "nota", None)
        calculada = getattr(r, "nota", None)
        estado = getattr(r, "estado_moodle", "") or ""
        retenido = getattr(r, "retenido_por", None)
        filas.append(
            [
                getattr(r, "alumno_nombre", None) or getattr(r, "alumno_idnumber", "") or "",
                getattr(r, "alumno_idnumber", "") or "",
                getattr(r, "alumno_email", "") or "",
                # Una celda vacia en la columna Nota se lee como un error del
                # archivo. Decir que todavia no hay nota es un dato.
                _celda_nota(nota, calculada, retenido),
                # El resultado sale RESUELTO del backend (`ResultadoNota`): antes
                # se decidia acá con ifs propios y una nota anulada salia
                # "Aprobado" mientras la pantalla decia "Anulada".
                etiqueta_resultado(getattr(r, "resultado", "") or ""),
                # Fuente única: el archivo y la pantalla dicen lo mismo. La
                # retención pisa al estado, igual que en el badge — si no, el
                # archivo decia "Falta conectar el campus" sobre una nota anulada.
                etiqueta_retencion(retenido) or etiqueta_estado_entrega(estado),
            ]
        )
    return filas
