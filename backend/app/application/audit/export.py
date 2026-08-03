"""Export del registro de auditoría a Excel y PDF.

Existe porque el registro de auditoría es JUSTAMENTE el que más sentido tiene poder
llevarse por período: ante un pedido de un organismo de control o una auditoría
interna (Ley 25.326), hace falta entregar "qué pasó entre estas dos fechas" en un
archivo, no una pantalla paginada. Estadísticas ya exportaba; esto no.

Respeta los MISMOS filtros que el listado (actor, acción, rango de fechas): lo que
se ve en pantalla es lo que sale en el archivo.

Sin jerga: las acciones salen con su etiqueta en castellano (``etiqueta_accion``),
nunca como ``moodle.sync``.
"""

from __future__ import annotations

import io
from datetime import datetime

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.application.stats.labels import (
    detalle_legible,
    etiqueta_accion,
    etiqueta_modulo,
)

_AZUL = "004BA8"
_HDR_FILL = PatternFill("solid", fgColor=_AZUL)
_HDR_FONT = Font(bold=True, color="FFFFFF")

# La IP NO se exporta: es dato sensible (Ley 25.326). El registro de auditoría se
# entrega a un organismo de control con "quién hizo qué y cuándo" — la dirección
# de red del actor no hace falta y su exposición es un riesgo de privacidad.
_COLUMNAS = ["Fecha y hora", "Usuario", "Módulo", "Acción", "Detalle"]
# Anchos pensados para que "Detalle" (el texto largo) no quede cortado en Excel.
_ANCHOS = [18, 40, 24, 34, 84]


def _usuario(e) -> str:
    """Identidad del actor para el reporte: 'Nombre (email)' — el email hace
    más seguimiento que el legajo — o el identificador solo si no hay nombre
    resuelto. NUNCA queda en blanco — un audit con la columna "Usuario" vacía
    no dice quién hizo la acción."""
    nombre = getattr(e, "actor_nombre", None)
    actor = getattr(e, "actor", "") or ""
    email = getattr(e, "actor_email", None)
    identificador = email if email and email != actor else actor
    if nombre and nombre != identificador:
        return f"{nombre} ({identificador})"
    return nombre or identificador


def _fecha_legible(iso: str | None) -> str:
    """ISO → 'dd/mm/aaaa hh:mm'. Si no parsea, devuelve el original (nunca vacío)."""
    if not iso:
        return ""
    try:
        return datetime.fromisoformat(iso.replace("Z", "+00:00")).strftime(
            "%d/%m/%Y %H:%M"
        )
    except (ValueError, TypeError):
        return iso


def _rango_texto(desde: str | None, hasta: str | None) -> str:
    if desde and hasta:
        return f"Período: {desde} a {hasta}"
    if desde:
        return f"Desde: {desde}"
    if hasta:
        return f"Hasta: {hasta}"
    return "Período: todo el registro"


def _filtros_texto(
    actor: str | None, accion: str | None, modulo: str | None = None
) -> str:
    partes = []
    if actor:
        partes.append(f"Usuario: {actor}")
    if modulo:
        partes.append(f"Módulo: {etiqueta_modulo(modulo)}")
    if accion:
        partes.append(f"Acción: {etiqueta_accion(accion)}")
    return " · ".join(partes) if partes else "Sin filtros de usuario, módulo ni acción"


def auditoria_a_xlsx(
    entradas: list,
    *,
    actor: str | None = None,
    accion: str | None = None,
    modulo: str | None = None,
    desde: str | None = None,
    hasta: str | None = None,
) -> bytes:
    """Registro de auditoría como .xlsx. Devuelve los bytes del archivo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoría"

    # Cabecera de contexto: un export sin el período ni los filtros aplicados no
    # se puede interpretar después (ni defender ante quien lo pidió).
    ws["A1"] = "Registro de auditoría — Active Exam"
    ws["A1"].font = Font(bold=True, size=14, color=_AZUL)
    ws["A2"] = _rango_texto(desde, hasta)
    ws["A3"] = _filtros_texto(actor, accion, modulo)
    ws["A4"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A5"] = f"Entradas exportadas: {len(entradas)}"

    fila_encabezado = 7
    for i, col in enumerate(_COLUMNAS, start=1):
        c = ws.cell(row=fila_encabezado, column=i, value=col)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="left")
        ws.column_dimensions[chr(64 + i)].width = _ANCHOS[i - 1]

    for j, e in enumerate(entradas, start=fila_encabezado + 1):
        ws.cell(row=j, column=1, value=_fecha_legible(getattr(e, "timestamp", None)))
        ws.cell(row=j, column=2, value=_usuario(e))
        ws.cell(row=j, column=3, value=etiqueta_modulo(getattr(e, "modulo", None)))
        ws.cell(row=j, column=4, value=etiqueta_accion(getattr(e, "accion", "") or ""))
        ws.cell(row=j, column=5, value=detalle_legible(getattr(e, "proposito", None)))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# FPDF con fuente core (Helvetica) solo soporta latin-1. Las rayas/puntos
# suspensivos/flechas tipográficas que usan labels.py y los propósitos NO
# entran en latin-1 y quedaban como "?" (BUG real: "Registro de auditoría ?
# Active Exam", "canjeando su contraseña ?", "Cambió ? Param: antes ?
# después"). Se transliteran a su equivalente ASCII ANTES del encode — Excel
# no tiene este problema (openpyxl es UTF-8 nativo), así que esto es
# exclusivo de este módulo de PDF.
_REEMPLAZOS_PDF_LATIN1 = {
    "—": "-",
    "–": "-",
    "…": "...",
    "→": "->",
    "«": '"',
    "»": '"',
    "‘": "'",
    "’": "'",
    "“": '"',
    "”": '"',
}


def _txt(valor: str) -> str:
    """FPDF core usa latin-1: translitera lo tipográfico común y reemplaza lo
    que aun así no entra (en vez de reventar)."""
    texto = valor or ""
    for feo, equivalente in _REEMPLAZOS_PDF_LATIN1.items():
        texto = texto.replace(feo, equivalente)
    return texto.encode("latin-1", "replace").decode("latin-1")


def auditoria_a_pdf(
    entradas: list,
    *,
    actor: str | None = None,
    accion: str | None = None,
    modulo: str | None = None,
    desde: str | None = None,
    hasta: str | None = None,
) -> bytes:
    """Registro de auditoría como PDF apaisado. Devuelve los bytes del archivo."""
    # Apaisado: la columna "Detalle" lleva frases completas y en vertical quedaban
    # ilegibles a fuerza de cortarlas.
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 15)
    pdf.set_text_color(0, 75, 168)
    pdf.cell(0, 9, _txt("Registro de auditoría — Active Exam"), new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 5, _txt(_rango_texto(desde, hasta)), new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, _txt(_filtros_texto(actor, accion, modulo)), new_x="LMARGIN", new_y="NEXT")
    pdf.cell(
        0,
        5,
        _txt(
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} · "
            f"Entradas: {len(entradas)}"
        ),
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.ln(3)

    # Sin IP: 5 columnas. Anchos suman ~273mm (A4 apaisado menos márgenes).
    anchos = [30, 55, 42, 52, 94]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(0, 75, 168)
    pdf.set_text_color(255, 255, 255)
    for ancho, col in zip(anchos, _COLUMNAS):
        pdf.cell(ancho, 8, _txt(col), border=0, fill=True, align="L")
    pdf.ln(8)

    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(30, 30, 30)
    for i, e in enumerate(entradas):
        if i % 2 == 0:
            pdf.set_fill_color(245, 247, 250)
            relleno = True
        else:
            pdf.set_fill_color(255, 255, 255)
            relleno = True
        valores = [
            _fecha_legible(getattr(e, "timestamp", None)),
            _usuario(e),
            etiqueta_modulo(getattr(e, "modulo", None)),
            etiqueta_accion(getattr(e, "accion", "") or ""),
            detalle_legible(getattr(e, "proposito", None)),
        ]
        for ancho, valor in zip(anchos, valores):
            # Recorte por ancho de columna: el PDF es una tabla, no un volcado.
            # El archivo completo sin recortes es el Excel.
            maximo = int(ancho / 1.7)
            texto = valor if len(valor) <= maximo else valor[: maximo - 3] + "..."
            pdf.cell(ancho, 6, _txt(texto), border=0, fill=relleno, align="L")
        pdf.ln(6)

    if not entradas:
        pdf.set_font("Helvetica", "I", 10)
        pdf.set_text_color(120, 120, 120)
        pdf.cell(
            0,
            8,
            _txt("No hay actividad registrada para los filtros aplicados."),
            new_x="LMARGIN",
            new_y="NEXT",
        )

    salida = pdf.output()
    return bytes(salida)
