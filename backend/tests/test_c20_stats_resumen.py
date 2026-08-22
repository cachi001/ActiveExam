"""C-20 (re-alcanzado): servicio de estadísticas institucionales standalone.

DB real (DATABASE_URL). Sin mocks de DB (regla dura). Verifica los conteos, las
personas en riesgo (score >= umbral) y la distribución de scores sobre datos que YA
existen — sin depender de C-13/C-16.
"""

from __future__ import annotations

import dataclasses
import json
import os
from datetime import UTC, datetime

import pytest
import pytest_asyncio
from fastapi import FastAPI
from httpx import ASGITransport, AsyncClient
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import NullPool

from app.application.stats.resumen_service import FiltrosStats, obtener_resumen
from app.infrastructure.auth.verifiers import encode_hs256
from app.infrastructure.persistence.base import Base
from app.infrastructure.persistence.models.exam_content import (  # noqa: F401
    ComisionModel,
    ExamenContenidoModel,
    MateriaModel,
)
from app.infrastructure.persistence.models.proctoring import (  # noqa: F401
    ProctoringEventModel,
    ProctoringSessionModel,
)
from app.infrastructure.persistence.models.transactional import (  # noqa: F401
    ConfiguracionSistemaModel,
    EventoScoreConfigModel,
)
from app.presentation.api.v1.stats.router import (
    ResumenStatsResponse,
    create_stats_router,
)
from tests.proctoring.conftest import (
    _TEST_JWT_AUDIENCE,
    _TEST_JWT_ISSUER,
    _TEST_JWT_SECRET,
    _build_test_jwt_validator,
)

UMBRAL = 40

_TABLES_TO_DROP = [
    "examen_contenido",
    "comision",
    "materia",
    "proctoring_event",
    "proctoring_biometria",
    "proctoring_session",
    "evento_score_config",
    "configuracion_sistema",
]
_TABLES_TO_CREATE = [
    MateriaModel.__table__,
    ComisionModel.__table__,
    ExamenContenidoModel.__table__,
    ProctoringSessionModel.__table__,
    ProctoringEventModel.__table__,
    ConfiguracionSistemaModel.__table__,
    EventoScoreConfigModel.__table__,
]


@pytest.fixture(scope="module")
def db_url():
    url = os.environ.get("DATABASE_URL")
    if not url:
        pytest.skip("DATABASE_URL no seteada — tests de integración omitidos")
    return url


@pytest_asyncio.fixture(scope="module")
async def engine(db_url):
    eng = create_async_engine(db_url, pool_pre_ping=True, future=True, poolclass=NullPool)
    async with eng.begin() as conn:
        for t in _TABLES_TO_DROP:
            await conn.execute(text(f'DROP TABLE IF EXISTS "{t}" CASCADE'))
        await conn.run_sync(Base.metadata.create_all, tables=_TABLES_TO_CREATE)
    yield eng
    async with eng.begin() as conn:
        for t in _TABLES_TO_DROP:
            await conn.execute(text(f'DROP TABLE IF EXISTS "{t}" CASCADE'))
    await eng.dispose()


@pytest_asyncio.fixture
async def session(engine) -> AsyncSession:
    factory = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)
    async with factory() as s:
        yield s
        await s.rollback()


async def _seed(s: AsyncSession) -> None:
    await s.execute(
        text(
            "TRUNCATE examen_contenido, comision, materia, proctoring_event, "
            "proctoring_session, configuracion_sistema, evento_score_config CASCADE"
        )
    )
    s.add(ConfiguracionSistemaModel(id="global", umbral_cola_revision=UMBRAL))
    s.add(
        EventoScoreConfigModel(
            tipo_evento="rostro_ausente", severidad="alta", peso=50, activo=True
        )
    )
    # 2 materias, 1 comisión, 2 exámenes
    m1 = MateriaModel(codigo="M1", nombre="Materia 1")
    m2 = MateriaModel(codigo="M2", nombre="Materia 2")
    s.add_all([m1, m2])
    await s.flush()
    s.add(
        ComisionModel(
            materia_id=m1.id, codigo="C1", nombre="Comisión 1",
            codigo_matriculacion="M1-C1",
        )
    )
    e1 = ExamenContenidoModel(titulo="E1")
    s.add_all([e1, ExamenContenidoModel(titulo="E2")])
    await s.flush()

    # sesión 1: finalizada + 1 evento rostro_ausente (score 50 >= 40 -> en riesgo)
    ses1 = ProctoringSessionModel(modo="examen", examen_contenido_id=e1.id)
    ses1.finalizada_en = datetime.now(UTC)
    # sesión 2: finalizada, sin eventos (score 0)
    ses2 = ProctoringSessionModel(modo="examen", examen_contenido_id=e1.id)
    ses2.finalizada_en = datetime.now(UTC)
    # sesión 3: NO finalizada, sin eventos
    ses3 = ProctoringSessionModel(modo="examen", examen_contenido_id=e1.id)
    s.add_all([ses1, ses2, ses3])
    await s.flush()
    s.add(
        ProctoringEventModel(
            session_id=ses1.id, tipo="rostro_ausente", severidad="alta",
            ts_cliente=datetime.now(UTC),
        )
    )
    await s.commit()


@pytest.mark.asyncio
async def test_resumen_conteos_y_riesgo(session):
    """Conteos correctos + 1 sesión en riesgo (score 50 >= umbral 40)."""
    await _seed(session)

    r = await obtener_resumen(session)

    assert r.total_materias == 2
    assert r.total_comisiones == 1
    assert r.total_examenes == 2
    assert r.total_sesiones == 3
    assert r.sesiones_finalizadas == 2
    assert r.umbral_riesgo == 40
    assert r.sesiones_en_riesgo == 1  # solo ses1 (score 50)


@pytest.mark.asyncio
async def test_resumen_distribucion_scores(session):
    """La distribución ubica ses1 en la banda de riesgo y ses2/ses3 en 0-24.

    Las bandas se derivan del UMBRAL vivo (acá 40), no de cortes fijos: la última
    arranca exactamente en el umbral. Antes las bandas eran 0-24/25-49/50-69/70-100
    pase lo que pase, así que con umbral 40 la sesión de score 50 caía en "50-69"
    mientras la banda "25-49" —que también contiene scores en riesgo (45)— no
    quedaba marcada. Ahora "banda alta" y "prioriza revisión" son lo mismo."""
    await _seed(session)

    r = await obtener_resumen(session)

    assert list(r.distribucion_scores.keys()) == ["0-24", "25-39", "40-100"]
    assert r.distribucion_scores["40-100"] == 1  # ses1 (score 50) — en riesgo
    assert r.distribucion_scores["0-24"] == 2
    assert r.distribucion_scores["25-39"] == 0
    # La banda alta coincide EXACTAMENTE con el conteo de sesiones en riesgo.
    assert r.distribucion_scores["40-100"] == r.sesiones_en_riesgo


@pytest.mark.asyncio
async def test_sesiones_de_diagnostico_no_cuentan(session):
    """Sesiones sin examen vinculado (Test de detección de Configuración,
    'Grabar sesión') NO deben sumar a ninguna métrica — no son un examen
    rendido. Mismo criterio que ya aplica la Cola de Revisión."""
    await _seed(session)
    diagnostico = ProctoringSessionModel(modo="examen", examen_contenido_id=None)
    diagnostico.finalizada_en = datetime.now(UTC)
    session.add(diagnostico)
    await session.flush()
    session.add(
        ProctoringEventModel(
            session_id=diagnostico.id, tipo="rostro_ausente", severidad="alta",
            ts_cliente=datetime.now(UTC),
        )
    )
    await session.commit()

    r = await obtener_resumen(session)

    # Sigue siendo 3 (ses1/ses2/ses3 del seed real) — la de diagnóstico no suma.
    assert r.total_sesiones == 3
    assert r.sesiones_finalizadas == 2
    assert r.sesiones_en_riesgo == 1


@pytest.mark.asyncio
async def test_resumen_vacio_da_ceros(session):
    """Sin datos: ceros legítimos (no error). Degradación segura."""
    await session.execute(
        text(
            "TRUNCATE examen_contenido, comision, materia, proctoring_event, "
            "proctoring_session, configuracion_sistema, evento_score_config CASCADE"
        )
    )
    await session.commit()

    r = await obtener_resumen(session)

    assert r.total_examenes == 0
    assert r.total_sesiones == 0
    assert r.sesiones_en_riesgo == 0


# ---------------------------------------------------------------------------
# Filtros + agregaciones nuevas (por materia / top eventos / por día / decisiones)
# ---------------------------------------------------------------------------


async def _seed_filtros(s: AsyncSession) -> dict[str, str]:
    """Semilla con vínculo real sesión→examen→comisión→materia para probar
    filtros y desgloses. Devuelve los ids relevantes."""
    await s.execute(
        text(
            "TRUNCATE examen_contenido, comision, materia, proctoring_event, "
            "proctoring_session, configuracion_sistema, evento_score_config CASCADE"
        )
    )
    s.add(ConfiguracionSistemaModel(id="global", umbral_cola_revision=UMBRAL))
    s.add_all([
        EventoScoreConfigModel(tipo_evento="rostro_ausente", severidad="alta", peso=50, activo=True),
        EventoScoreConfigModel(tipo_evento="cambio_pestana", severidad="media", peso=20, activo=True),
    ])
    m1 = MateriaModel(codigo="ALG", nombre="Álgebra")
    m2 = MateriaModel(codigo="FIS", nombre="Física")
    s.add_all([m1, m2])
    await s.flush()
    c1 = ComisionModel(materia_id=m1.id, codigo="C1", nombre="Com 1", codigo_matriculacion="ALG-C1")
    c2 = ComisionModel(materia_id=m2.id, codigo="C1", nombre="Com 1", codigo_matriculacion="FIS-C1")
    s.add_all([c1, c2])
    await s.flush()
    e1 = ExamenContenidoModel(titulo="Parcial Álgebra", comision_id=c1.id)
    e2 = ExamenContenidoModel(titulo="Parcial Física", comision_id=c2.id)
    s.add_all([e1, e2])
    await s.flush()

    dia1 = datetime(2026, 7, 1, 10, 0, tzinfo=UTC)
    dia2 = datetime(2026, 7, 2, 10, 0, tzinfo=UTC)
    # Materia 1 (Álgebra): 2 sesiones — sa1 en riesgo (rostro_ausente=50>=40) + revisada.
    sa1 = ProctoringSessionModel(modo="examen", examen_contenido_id=e1.id)
    sa1.creada_en = dia1
    sa1.finalizada_en = dia1
    sa1.decision = "anulado"
    sa2 = ProctoringSessionModel(modo="examen", examen_contenido_id=e1.id)
    sa2.creada_en = dia2
    sa2.finalizada_en = dia2
    # Materia 2 (Física): 1 sesión — cambio_pestana=20 < 40 (no riesgo).
    sf1 = ProctoringSessionModel(modo="examen", examen_contenido_id=e2.id)
    sf1.creada_en = dia2
    sf1.finalizada_en = dia2
    s.add_all([sa1, sa2, sf1])
    await s.flush()
    s.add_all([
        ProctoringEventModel(session_id=sa1.id, tipo="rostro_ausente", severidad="alta", ts_cliente=dia1),
        ProctoringEventModel(session_id=sf1.id, tipo="cambio_pestana", severidad="media", ts_cliente=dia2),
    ])
    await s.commit()
    return {"m1": m1.id, "m2": m2.id, "e1": e1.id, "e2": e2.id}


@pytest.mark.asyncio
async def test_filtro_por_materia_acota_sesiones(session):
    ids = await _seed_filtros(session)

    todos = await obtener_resumen(session)
    assert todos.total_sesiones == 3  # sin filtro: las 3

    solo_alg = await obtener_resumen(session, FiltrosStats(materia_id=ids["m1"]))
    assert solo_alg.total_sesiones == 2  # solo las 2 de Álgebra
    assert solo_alg.sesiones_en_riesgo == 1  # sa1
    # El desglose por materia trae SOLO Álgebra al filtrar.
    assert [m.nombre for m in solo_alg.por_materia] == ["Álgebra"]
    assert solo_alg.por_materia[0].sesiones == 2
    assert solo_alg.por_materia[0].en_riesgo == 1


@pytest.mark.asyncio
async def test_por_materia_sin_filtro_lista_todas(session):
    await _seed_filtros(session)
    r = await obtener_resumen(session)
    # Ordenado por sesiones desc: Álgebra (2) antes que Física (1).
    assert [m.nombre for m in r.por_materia] == ["Álgebra", "Física"]
    assert r.por_materia[0].sesiones == 2
    assert r.por_materia[1].sesiones == 1


@pytest.mark.asyncio
async def test_top_eventos_ordena_por_frecuencia(session):
    await _seed_filtros(session)
    r = await obtener_resumen(session)
    tipos = {e.tipo: e.cantidad for e in r.top_eventos}
    assert tipos == {"rostro_ausente": 1, "cambio_pestana": 1}


@pytest.mark.asyncio
async def test_por_dia_y_decisiones(session):
    await _seed_filtros(session)
    r = await obtener_resumen(session)
    por_dia = {d.fecha: d.sesiones for d in r.por_dia}
    assert por_dia == {"2026-07-01": 1, "2026-07-02": 2}
    # `decisiones` describe la COLA DE REVISIÓN, no el padrón entero: solo entran
    # las sesiones con score >= umbral (las que efectivamente van a revisión
    # humana). Las que nunca la necesitaron no cuentan como "sin revisar" — si no,
    # el donut de "Estado de revisión" quedaría dominado por sesiones limpias que
    # nadie tiene que mirar. Este test pedía las 3 sesiones; el servicio acota a la
    # cola desde que se documentó ese criterio.
    assert r.decisiones.get("anulado") == 1
    assert r.decisiones.get("sin_revisar") is None
    assert sum(r.decisiones.values()) == r.sesiones_en_riesgo


@pytest.mark.asyncio
async def test_filtro_materia_id_invalido_no_rompe(session):
    """Un materia_id malformado (no-UUID) filtra a vacío, no tira 500."""
    await _seed_filtros(session)
    r = await obtener_resumen(session, FiltrosStats(materia_id="no-es-uuid"))
    assert r.total_sesiones == 0
    assert r.por_materia == []


@pytest.mark.asyncio
async def test_filtro_rango_fechas(session):
    await _seed_filtros(session)
    # Solo el 2026-07-02 → 2 sesiones (sa2 + sf1).
    r = await obtener_resumen(
        session, FiltrosStats(desde="2026-07-02T00:00:00+00:00")
    )
    assert r.total_sesiones == 2


# ---------------------------------------------------------------------------
# Endpoint GET /api/v1/stats/resumen (RBAC admin_sistema/coordinador)
# ---------------------------------------------------------------------------


def _token(roles) -> str:
    claims = {
        "iss": _TEST_JWT_ISSUER,
        "aud": _TEST_JWT_AUDIENCE,
        "sub": "sub-stats",
        "preferred_username": "u-stats",
        "email": "s@u.edu",
        "exp": 9999999999,
        "amr": ["otp"],
        "realm_access": {"roles": list(roles)},
    }
    return encode_hs256(claims, _TEST_JWT_SECRET)


@pytest_asyncio.fixture
async def app_stats(engine):
    factory = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)
    async with factory() as s:
        await _seed(s)
    app = FastAPI()
    app.state.jwt_validator = _build_test_jwt_validator()
    app.include_router(
        create_stats_router(session_factory=factory), prefix="/api/v1/stats"
    )
    return app


@pytest.mark.asyncio
async def test_endpoint_resumen_admin_200(app_stats):
    async with AsyncClient(
        transport=ASGITransport(app=app_stats),
        base_url="http://test",
        headers={"Authorization": f"Bearer {_token(['admin_sistema'])}"},
    ) as c:
        resp = await c.get("/api/v1/stats/resumen")
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["total_materias"] == 2
    assert body["sesiones_en_riesgo"] == 1
    assert body["umbral_riesgo"] == 40
    assert "distribucion_scores" in body


@pytest.mark.asyncio
async def test_endpoint_resumen_estudiante_403(app_stats):
    async with AsyncClient(
        transport=ASGITransport(app=app_stats),
        base_url="http://test",
        headers={"Authorization": f"Bearer {_token(['estudiante'])}"},
    ) as c:
        resp = await c.get("/api/v1/stats/resumen")
    assert resp.status_code == 403, resp.text


@pytest.mark.asyncio
async def test_endpoint_resumen_tutor_403(app_stats):
    """c-79: el tutor NO ve estadísticas institucionales — los filtros de este
    endpoint son query params libres, sin scoping por pertenencia, así que
    darle acceso permitiría pedir el resumen de una comisión ajena."""
    async with AsyncClient(
        transport=ASGITransport(app=app_stats),
        base_url="http://test",
        headers={"Authorization": f"Bearer {_token(['tutor'])}"},
    ) as c:
        resp = await c.get("/api/v1/stats/resumen")
    assert resp.status_code == 403, resp.text


@pytest.mark.asyncio
async def test_endpoint_resumen_coordinador_200(app_stats):
    async with AsyncClient(
        transport=ASGITransport(app=app_stats),
        base_url="http://test",
        headers={"Authorization": f"Bearer {_token(['coordinador'])}"},
    ) as c:
        resp = await c.get("/api/v1/stats/resumen")
    assert resp.status_code == 200, resp.text


@pytest.mark.asyncio
async def test_endpoint_resumen_incluye_agregaciones_nuevas(app_stats):
    async with AsyncClient(
        transport=ASGITransport(app=app_stats),
        base_url="http://test",
        headers={"Authorization": f"Bearer {_token(['admin_sistema'])}"},
    ) as c:
        resp = await c.get("/api/v1/stats/resumen")
    assert resp.status_code == 200, resp.text
    body = resp.json()
    for clave in ("por_materia", "top_eventos", "por_dia", "decisiones"):
        assert clave in body


@pytest.mark.asyncio
async def test_endpoint_export_pdf(app_stats):
    async with AsyncClient(
        transport=ASGITransport(app=app_stats),
        base_url="http://test",
        headers={"Authorization": f"Bearer {_token(['admin_sistema'])}"},
    ) as c:
        resp = await c.get("/api/v1/stats/export.pdf")
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/pdf"
    assert "attachment" in resp.headers.get("content-disposition", "")
    # Un PDF válido empieza con el magic "%PDF".
    assert resp.content[:4] == b"%PDF"
    # El PDF embebe el dashboard de gráficos (XObject de imagen).
    assert b"/Image" in resp.content


@pytest.mark.asyncio
async def test_endpoint_export_xlsx(app_stats):
    import io as _io

    from openpyxl import load_workbook

    async with AsyncClient(
        transport=ASGITransport(app=app_stats),
        base_url="http://test",
        headers={"Authorization": f"Bearer {_token(['admin_sistema'])}"},
    ) as c:
        resp = await c.get("/api/v1/stats/export.xlsx")
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "attachment" in resp.headers.get("content-disposition", "")
    # Un .xlsx es un ZIP → magic "PK".
    assert resp.content[:2] == b"PK"
    # Es un Excel real: se abre, tiene las hojas y al menos un gráfico nativo.
    wb = load_workbook(_io.BytesIO(resp.content))
    assert "Resumen" in wb.sheetnames
    assert "Por materia" in wb.sheetnames
    # Los gráficos NO viven en "Resumen" (que es la tabla de totales): cada hoja de
    # datos lleva el suyo cuando tiene filas, y "Panel" abre con el dashboard
    # completo como imagen. El test pedía un gráfico en "Resumen" desde antes de esa
    # reorganización. Se afirma sobre el libro entero para no atarse a qué hojas
    # tienen datos en el fixture (una hoja vacía escribe "Sin datos", sin gráfico).
    assert "Panel" in wb.sheetnames
    assert len(wb["Panel"]._images) >= 1
    assert sum(len(wb[h]._charts) for h in wb.sheetnames) >= 1


# ---------------------------------------------------------------------------
# Gobernanza L2.5 + privacidad (Ley 25.326): invariancia, no-veredicto, PII.
# Términos prohibidos = cualquier campo que implique un JUICIO automático.
# ---------------------------------------------------------------------------

_TERMINOS_VEREDICTO = {
    "veredicto",
    "sancion",
    "sanción",
    "culpable",
    "acusacion",
    "acusación",
    "decision_automatica",
    "penalizacion",
    "castigo",
    "fraude",
}
# Claves que delatarían PII a nivel de individuo en un agregado institucional.
_CLAVES_PII = {
    "email",
    "dni",
    "legajo",
    "nombre_estudiante",
    "apellido",
    "estudiante",
    "alumno",
    "usuario_id",
    "sub",
}

_TABLAS_INVARIANZA = [
    "examen_contenido",
    "comision",
    "materia",
    "proctoring_session",
    "proctoring_event",
    "configuracion_sistema",
    "evento_score_config",
]


def _claves_recursivas(obj) -> set[str]:
    """Todas las claves (en minúscula) que aparecen en un JSON anidado."""
    claves: set[str] = set()
    if isinstance(obj, dict):
        for k, v in obj.items():
            claves.add(k.lower())
            claves |= _claves_recursivas(v)
    elif isinstance(obj, list):
        for v in obj:
            claves |= _claves_recursivas(v)
    return claves


@pytest.mark.asyncio
async def test_capa_no_muta_nada_invariancia(session):
    """Task 1.4: la capa SOLO lee — llamarla no cambia ninguna fila."""
    await _seed(session)

    async def _snapshot() -> dict[str, int]:
        conteos = {}
        for t in _TABLAS_INVARIANZA:
            conteos[t] = int(
                (await session.execute(text(f'SELECT count(*) FROM "{t}"'))).scalar_one()
            )
        return conteos

    antes = await _snapshot()
    await obtener_resumen(session)
    await obtener_resumen(session, FiltrosStats(materia_id="no-es-uuid"))
    despues = await _snapshot()

    assert antes == despues  # ni un INSERT/UPDATE/DELETE


@pytest.mark.asyncio
async def test_contrato_riesgo_es_senal_no_veredicto(session):
    """Task 2.4: el 'riesgo' es un CONTEO agregado (int), nunca un veredicto.

    El dataclass del sumario no tiene ningún campo de juicio/sanción."""
    await _seed(session)
    r = await obtener_resumen(session)

    assert isinstance(r.sesiones_en_riesgo, int)  # conteo, no booleano de culpa
    assert isinstance(r.umbral_riesgo, int)
    campos = {f.name for f in dataclasses.fields(r)}
    assert _TERMINOS_VEREDICTO.isdisjoint(campos), campos


@pytest.mark.asyncio
async def test_endpoint_resumen_no_expone_veredicto(app_stats):
    """Task 5.1 (transversal): ningún path del informe emite veredicto/acción."""
    async with AsyncClient(
        transport=ASGITransport(app=app_stats),
        base_url="http://test",
        headers={"Authorization": f"Bearer {_token(['admin_sistema'])}"},
    ) as c:
        resp = await c.get("/api/v1/stats/resumen")
    assert resp.status_code == 200, resp.text
    body = resp.json()

    claves = _claves_recursivas(body)
    assert _TERMINOS_VEREDICTO.isdisjoint(claves), claves
    # El riesgo llega como conteo entero (señal de priorización), no como fallo.
    assert isinstance(body["sesiones_en_riesgo"], int)


@pytest.mark.asyncio
async def test_endpoint_resumen_sin_pii_y_forbid(app_stats):
    """Task 3.2: el agregado no expone PII y el schema rechaza campos extra."""
    # El schema de salida es un allowlist estricto (extra='forbid').
    assert ResumenStatsResponse.model_config.get("extra") == "forbid"

    async with AsyncClient(
        transport=ASGITransport(app=app_stats),
        base_url="http://test",
        headers={"Authorization": f"Bearer {_token(['admin_sistema'])}"},
    ) as c:
        resp = await c.get("/api/v1/stats/resumen")
    assert resp.status_code == 200, resp.text
    body = resp.json()

    # Ningún email en el agregado (marcador barato pero efectivo de PII).
    assert "@" not in json.dumps(body, ensure_ascii=False)
    assert _CLAVES_PII.isdisjoint(_claves_recursivas(body))


@pytest.mark.asyncio
async def test_minimizacion_pii_en_desgloses(session):
    """Task 5.2 (transversal): aun con sesiones sembradas, los desgloses agregan
    y NO exponen identidad individual — solo id de catálogo + conteos."""
    await _seed_filtros(session)
    r = await obtener_resumen(session)

    assert r.por_materia  # hay desglose que revisar
    for m in r.por_materia:
        assert {f.name for f in dataclasses.fields(m)} == {
            "materia_id",
            "nombre",
            "sesiones",
            "en_riesgo",
        }
    for e in r.top_eventos:
        assert {f.name for f in dataclasses.fields(e)} == {"tipo", "cantidad"}
    for d in r.por_dia:
        assert {f.name for f in dataclasses.fields(d)} == {"fecha", "sesiones"}
