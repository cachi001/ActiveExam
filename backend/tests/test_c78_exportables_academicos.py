"""c-78 §13 (E-10, D14) — exports académicos y marcado manual de la nota.

Cubre contra DB REAL:
- export de inscriptos de una comisión en Excel y PDF, con las columnas acordadas
  para cruzar contra Moodle;
- export de notas del examen;
- paginación del listado de inscriptos;
- marcado manual de la nota: registra quién y cuándo, produce un estado
  distinguible de 'enviado', y NO puede pisar una confirmación del campus.
"""

from __future__ import annotations

import io
import os
import uuid
from datetime import UTC, datetime

import pytest
import pytest_asyncio
from fastapi import FastAPI
from httpx import ASGITransport, AsyncClient
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import NullPool

from app.application.exam_content.export import (
    COLUMNAS_INSCRIPTOS,
    COLUMNAS_NOTAS,
    filas_inscriptos,
    filas_notas,
)
from app.infrastructure.persistence.base import Base
from app.infrastructure.persistence.models.comision_tutor import (  # noqa: F401
    ComisionTutorModel,
    MateriaCoordinadorModel,
    MateriaProfesorModel,
)
from app.infrastructure.persistence.models.exam_content import (  # noqa: F401
    ComisionModel,
    ExamenContenidoModel,
    MateriaModel,
    OpcionRespuestaModel,
    PreguntaExamenModel,
)
from app.infrastructure.persistence.models.inscripcion import InscripcionModel  # noqa: F401
from app.infrastructure.persistence.models.moodle_writeback import (  # noqa: F401
    MoodleWritebackAuditModel,
    MoodleWritebackEstadoModel,
    RespuestaAlumnoModel,
)
from app.infrastructure.persistence.models.proctoring import (  # noqa: F401
    ProctoringEventModel,
    ProctoringSessionModel,
)
from app.infrastructure.persistence.models.transactional import UsuarioModel  # noqa: F401
from app.presentation.api.v1.exam_content.router import create_exam_content_router
from tests.proctoring.conftest import _build_test_jwt_validator, auth_headers

pytestmark = pytest.mark.asyncio

_TABLES_TO_DROP = [
    "moodle_writeback_audit",
    "moodle_writeback_estado",
    "respuesta_alumno",
    "opcion_respuesta",
    "pregunta_examen",
    "proctoring_event",
    "proctoring_session",
    "examen_contenido",
    "inscripcion",
    "comision_tutor",
    "materia_coordinador",
    "materia_profesor",
    "comision",
    "materia",
]
_TABLES_TO_CREATE = [
    UsuarioModel.__table__,
    MateriaModel.__table__,
    MateriaCoordinadorModel.__table__,
    MateriaProfesorModel.__table__,
    ComisionModel.__table__,
    ComisionTutorModel.__table__,
    InscripcionModel.__table__,
    ExamenContenidoModel.__table__,
    PreguntaExamenModel.__table__,
    OpcionRespuestaModel.__table__,
    ProctoringSessionModel.__table__,
    ProctoringEventModel.__table__,
    RespuestaAlumnoModel.__table__,
    MoodleWritebackEstadoModel.__table__,
    MoodleWritebackAuditModel.__table__,
]


# ===========================================================================
# Bloque A — proyección a filas (PURO, sin DB)
# ===========================================================================


class _Inscripto:
    def __init__(self, **kw):
        self.__dict__.update(kw)


def test_las_filas_de_inscriptos_tienen_las_columnas_acordadas():
    """Las 5 de identificación para cruzar contra Moodle, MÁS la elegibilidad.

    Las tres últimas se agregaron porque el archivo se descarga justamente para
    saber quién no va a poder rendir (sin consentimiento o sin biometría no se
    puede), y ese dato quedaba solo en la pantalla, alumno por alumno.
    """
    assert [c.titulo for c in COLUMNAS_INSCRIPTOS] == [
        "Apellido",
        "Nombre",
        "Usuario",
        "Email",
        "Inscripción",
        "Consentimiento",
        "Biometría",
        "¿Puede rendir?",
    ]
    filas = filas_inscriptos(
        [
            _Inscripto(
                apellido="Pérez",
                nombre="Juana",
                username="EST-001",
                email="juana@uni.edu",
                inscripto_en=datetime(2026, 3, 15, 9, 30, tzinfo=UTC),
                consentimiento_vigente=True,
                biometria_vigente=True,
                puede_rendir=True,
            )
        ]
    )
    assert filas == [
        [
            "Pérez",
            "Juana",
            "EST-001",
            "juana@uni.edu",
            "15/03/2026 09:30",
            "Sí",
            "Sí",
            "Sí",
        ]
    ]


def test_un_inscripto_sin_nombre_no_saca_none_en_el_archivo():
    """Un 'None' impreso en un listado que se cruza a mano es basura, no un dato.

    Sin datos de elegibilidad se informa "No"/"NO", que es la lectura conservadora:
    decir que sí puede rendir sin saberlo mandaría al alumno a un examen que el
    sistema le va a bloquear.
    """
    filas = filas_inscriptos(
        [_Inscripto(apellido=None, nombre=None, username="EST-002", email="x@uni.edu")]
    )
    assert filas == [["", "", "EST-002", "x@uni.edu", "", "No", "No", "NO"]]


def test_las_notas_salen_con_el_estado_en_castellano():
    """El archivo lo lee quien carga notas a mano: 'sin_token' no le dice nada."""
    filas = filas_notas(
        [
            _Inscripto(
                alumno_nombre="Juana Pérez",
                alumno_idnumber="EST-001",
                alumno_email="juana@uni.edu",
                nota=8.5,
                nota_efectiva=8.5,
                # El resultado lo RESUELVE el backend y el export solo lo
                # muestra: antes lo decidía acá con ifs propios y una nota
                # anulada salía "Aprobado" mientras la pantalla decía "Anulada".
                resultado="sin_criterio",
                estado_entrega="finalizada",
                estado_moodle="pendiente",
            ),
            _Inscripto(
                alumno_nombre=None,
                alumno_idnumber="EST-002",
                alumno_email="x@uni.edu",
                nota=None,
                nota_efectiva=None,
                resultado="sin_nota",
                estado_entrega="no_finalizada",
                estado_moodle="manual",
            ),
        ]
    )
    assert filas[0] == [
        "Juana Pérez",
        "EST-001",
        "juana@uni.edu",
        "8.50",
        # Dos columnas para dos cosas distintas: si aprobó, y si la nota se
        # entregó al campus. La tercera ("Examen del alumno") se sacó: decía lo
        # mismo que la de la entrega con otras palabras y confundía cuál era cuál.
        "Sin criterio de aprobación",
        # La etiqueta sale del enum `EstadoEntregaNota` (fuente única): el archivo dice
        # exactamente lo mismo que la pantalla. Tenían textos distintos para el
        # mismo estado.
        "Pendiente",
    ]
    # Sin nota: celda vacía, no "None" ni "0".
    # Una celda vacia en la columna Nota se lee como un error del archivo; decir
    # que todavia no hay nota es un dato, y ademas la distingue de un cero.
    assert filas[1][3] == "Sin nota"
    assert filas[1][0] == "EST-002", "sin nombre cargado cae al legajo, no a un UUID"
    assert filas[1][4] == "Sin nota", "sin nota no es desaprobado"
    assert filas[1][5] == "Cargada a mano"


def test_las_columnas_de_notas_no_exponen_proctoring():
    """PRIVACIDAD: el export lleva lo mínimo para su propósito declarado."""
    titulos = " ".join(c.titulo.lower() for c in COLUMNAS_NOTAS)
    for prohibido in ("score", "riesgo", "evento", "evidencia", "captura"):
        assert prohibido not in titulos, f"el export no debe exponer {prohibido}"


# ===========================================================================
# Bloque B — endpoints (DB real)
# ===========================================================================


@pytest.fixture(scope="module")
def db_url():
    url = os.environ.get("DATABASE_URL")
    if not url:
        pytest.skip("DATABASE_URL no seteada — tests de integración omitidos")
    return url


@pytest_asyncio.fixture(scope="module")
async def engine(db_url):
    eng = create_async_engine(db_url, pool_pre_ping=True, future=True, poolclass=NullPool)
    async with eng.begin() as conn:
        for t in _TABLES_TO_DROP:
            await conn.execute(text(f'DROP TABLE IF EXISTS "{t}" CASCADE'))
        await conn.run_sync(Base.metadata.create_all, tables=_TABLES_TO_CREATE)
    yield eng
    async with eng.begin() as conn:
        for t in _TABLES_TO_DROP:
            await conn.execute(text(f'DROP TABLE IF EXISTS "{t}" CASCADE'))
    await eng.dispose()


@pytest_asyncio.fixture
async def factory(engine):
    return async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)


@pytest_asyncio.fixture
async def app(factory):
    application = FastAPI()
    application.state.jwt_validator = _build_test_jwt_validator()
    application.state.session_factory = factory
    application.include_router(
        create_exam_content_router(session_factory=factory, writeback_svc=None),
        prefix="/api/v1/exam-content",
    )
    return application


def _admin(app):
    return AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://test",
        headers=auth_headers(["admin_sistema"], subject="staff-export"),
    )


def _fila_encabezado(ws, primer_titulo: str) -> int:
    """En que fila arranca la tabla.

    No se puede fijar en 6: arriba de la tabla va un bloque de resumen (cuantos
    pueden rendir, cuantos aprobaron) cuyo largo depende de los datos, porque las
    lineas en cero no se imprimen. Buscar el encabezado por su primer titulo deja
    el test midiendo el contenido y no el layout.
    """
    for fila in range(1, 30):
        if ws.cell(row=fila, column=1).value == primer_titulo:
            return fila
    raise AssertionError(f"No se encontro el encabezado '{primer_titulo}' en el archivo")



async def _comision_con_inscriptos(factory, cantidad: int) -> tuple[str, str]:
    """Materia + comisión + N alumnos inscriptos. Devuelve (comision_id, examen_id)."""
    sufijo = uuid.uuid4().hex[:6]
    async with factory() as s:
        materia = MateriaModel(codigo=f"MAT-{sufijo}", nombre=f"Materia {sufijo}")
        s.add(materia)
        await s.flush()
        comision = ComisionModel(
            materia_id=materia.id,
            codigo=f"C-{sufijo}",
            nombre=f"Comisión {sufijo}",
            codigo_matriculacion=f"K-{sufijo}",
            periodo="1C",
            anio=2026,
        )
        s.add(comision)
        await s.flush()
        for i in range(cantidad):
            alumno = UsuarioModel(
                username=f"EST-{sufijo}-{i:02d}",
                email=f"est{i}-{sufijo}@uni.edu",
                nombre=f"Alumno{i}",
                apellido=f"Apellido{i:02d}",
                roles=["estudiante"],
            )
            s.add(alumno)
            await s.flush()
            s.add(InscripcionModel(usuario_id=alumno.id, comision_id=comision.id))
        examen = ExamenContenidoModel(titulo=f"Parcial {sufijo}", comision_id=comision.id)
        s.add(examen)
        await s.flush()
        ids = (comision.id, examen.id)
        await s.commit()
    return ids


async def _resultado_con_nota(factory, examen_id: str, *, estado: str) -> str:
    """Sesión finalizada + fila de write-back en el estado dado. Devuelve session_id."""
    ahora = datetime.now(UTC)
    async with factory() as s:
        sesion = ProctoringSessionModel(
            modo="examen",
            examen_contenido_id=examen_id,
            alumno_idnumber="EST-NOTA",
            alumno_email="est-nota@uni.edu",
            creada_en=ahora,
            finalizada_en=ahora,
        )
        s.add(sesion)
        await s.flush()
        s.add(
            MoodleWritebackEstadoModel(
                session_id=sesion.id,
                alumno_idnumber="EST-NOTA",
                alumno_email="est-nota@uni.edu",
                nota=7.5,
                estado=estado,
            )
        )
        sid = sesion.id
        await s.commit()
    return sid


# --- §13.2 paginación -----------------------------------------------------


async def test_el_listado_de_inscriptos_se_pagina(app, factory):
    comision_id, _e = await _comision_con_inscriptos(factory, 12)

    async with _admin(app) as c:
        pagina = await c.get(
            f"/api/v1/exam-content/comisiones/{comision_id}/alumnos",
            params={"page": 1, "page_size": 5},
        )
        segunda = await c.get(
            f"/api/v1/exam-content/comisiones/{comision_id}/alumnos",
            params={"page": 2, "page_size": 5},
        )

    assert pagina.status_code == 200, pagina.text
    cuerpo = pagina.json()
    assert len(cuerpo["items"]) == 5
    assert cuerpo["total"] == 12, "`total` cuenta a TODOS los inscriptos, no la página"
    assert cuerpo["page"] == 1

    ids_p1 = {i["usuario_id"] for i in cuerpo["items"]}
    ids_p2 = {i["usuario_id"] for i in segunda.json()["items"]}
    assert not (ids_p1 & ids_p2), "las páginas no se pueden solapar"


async def test_un_page_size_grande_devuelve_todo(app, factory):
    """Quien quiera la lista entera la sigue pudiendo pedir."""
    comision_id, _e = await _comision_con_inscriptos(factory, 12)

    async with _admin(app) as c:
        resp = await c.get(
            f"/api/v1/exam-content/comisiones/{comision_id}/alumnos",
            params={"page_size": 500},
        )

    assert len(resp.json()["items"]) == 12


# --- §13.4 export de inscriptos -------------------------------------------


async def test_export_de_inscriptos_a_excel(app, factory):
    from openpyxl import load_workbook

    comision_id, _e = await _comision_con_inscriptos(factory, 3)

    async with _admin(app) as c:
        resp = await c.get(
            f"/api/v1/exam-content/comisiones/{comision_id}/alumnos/export.xlsx"
        )

    assert resp.status_code == 200, resp.text
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    hdr = _fila_encabezado(ws, "Apellido")
    encabezados = [ws.cell(row=hdr, column=i).value for i in range(1, 6)]
    assert encabezados == ["Apellido", "Nombre", "Usuario", "Email", "Inscripción"]
    # 3 inscriptos → 3 filas de datos a partir de la 7.
    assert ws.cell(row=hdr + 1, column=1).value == "Apellido00"
    assert ws.cell(row=hdr + 3, column=1).value == "Apellido02"


async def test_export_de_inscriptos_a_pdf(app, factory):
    comision_id, _e = await _comision_con_inscriptos(factory, 3)

    async with _admin(app) as c:
        resp = await c.get(
            f"/api/v1/exam-content/comisiones/{comision_id}/alumnos/export.pdf"
        )

    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content.startswith(b"%PDF"), "no es un PDF válido"


async def test_el_export_de_inscriptos_exige_pertenencia(app, factory):
    comision_id, _e = await _comision_con_inscriptos(factory, 2)

    ajeno = AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://test",
        headers=auth_headers(["tutor"], subject=str(uuid.uuid4())),
    )
    async with ajeno as c:
        resp = await c.get(
            f"/api/v1/exam-content/comisiones/{comision_id}/alumnos/export.xlsx"
        )

    assert resp.status_code == 403, resp.text


# --- §13.5 export de notas -------------------------------------------------


async def test_export_de_notas_del_examen(app, factory):
    from openpyxl import load_workbook

    _c, examen_id = await _comision_con_inscriptos(factory, 1)
    await _resultado_con_nota(factory, examen_id, estado="pendiente")

    async with _admin(app) as c:
        xlsx = await c.get(f"/api/v1/exam-content/{examen_id}/notas/export.xlsx")
        pdf = await c.get(f"/api/v1/exam-content/{examen_id}/notas/export.pdf")

    assert xlsx.status_code == 200, xlsx.text
    assert pdf.status_code == 200, pdf.text
    assert pdf.content.startswith(b"%PDF")

    ws = load_workbook(io.BytesIO(xlsx.content)).active
    hdr = _fila_encabezado(ws, "Alumno")
    assert [ws.cell(row=hdr, column=i).value for i in range(1, 7)] == [
        "Alumno",
        "Usuario",
        "Email",
        "Nota",
        # "Resultado" es aprobado/desaprobado; "Estado de la entrega" es si la
        # nota llegó al campus. Son dos cosas distintas y no se mezclan.
        "Resultado",
        "Estado de la entrega",
    ]
    assert ws.cell(row=hdr + 1, column=4).value == "7.50"


# --- §13.6 marcado manual (D14) -------------------------------------------


async def test_marcar_a_mano_registra_quien_y_cuando(app, factory):
    _c, examen_id = await _comision_con_inscriptos(factory, 1)
    session_id = await _resultado_con_nota(factory, examen_id, estado="pendiente")

    async with _admin(app) as c:
        resp = await c.patch(
            f"/api/v1/exam-content/{examen_id}/resultados/{session_id}/marcar-cargada"
        )

    assert resp.status_code == 200, resp.text
    cuerpo = resp.json()
    assert cuerpo["estado_moodle"] == "manual", (
        "el estado manual tiene que ser DISTINGUIBLE de 'enviado'"
    )
    assert cuerpo["marcada_manual_por"], "tiene que quedar QUIÉN lo marcó"
    assert cuerpo["marcada_manual_en"], "tiene que quedar CUÁNDO"

    async with factory() as s:
        fila = (
            await s.execute(
                select(MoodleWritebackEstadoModel).where(
                    MoodleWritebackEstadoModel.session_id == session_id
                )
            )
        ).scalar_one()
        assert fila.estado == "manual"
        assert fila.marcada_manual_por is not None
        assert fila.marcada_manual_en is not None


async def test_no_se_puede_pisar_una_confirmacion_del_campus(app, factory):
    """D14: una afirmación humana no reemplaza una confirmación del sistema."""
    _c, examen_id = await _comision_con_inscriptos(factory, 1)
    session_id = await _resultado_con_nota(factory, examen_id, estado="enviado")

    async with _admin(app) as c:
        resp = await c.patch(
            f"/api/v1/exam-content/{examen_id}/resultados/{session_id}/marcar-cargada"
        )

    assert resp.status_code == 409, resp.text
    assert resp.json()["detail"]["error"] == "nota_confirmada_por_el_campus"

    async with factory() as s:
        fila = (
            await s.execute(
                select(MoodleWritebackEstadoModel).where(
                    MoodleWritebackEstadoModel.session_id == session_id
                )
            )
        ).scalar_one()
        assert fila.estado == "enviado", "el rechazo no puede haber cambiado el estado"
        assert fila.marcada_manual_por is None


async def test_marcar_una_sesion_de_otro_examen_da_404(app, factory):
    """El session_id suelto no alcanza: tiene que ser de ESTE examen."""
    _c1, examen_a = await _comision_con_inscriptos(factory, 1)
    _c2, examen_b = await _comision_con_inscriptos(factory, 1)
    session_id = await _resultado_con_nota(factory, examen_a, estado="pendiente")

    async with _admin(app) as c:
        resp = await c.patch(
            f"/api/v1/exam-content/{examen_b}/resultados/{session_id}/marcar-cargada"
        )

    assert resp.status_code == 404, resp.text
